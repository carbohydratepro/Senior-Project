from transformers import BertJapaneseTokenizer, BertModel
import torch
import numpy as np
import pandas as pd
import openpyxl
import statistics


class SentenceBertJapanese:
    def __init__(self, model_name_or_path, device=None):
        self.tokenizer = BertJapaneseTokenizer.from_pretrained(model_name_or_path)
        self.model = BertModel.from_pretrained(model_name_or_path)
        self.model.eval()

        if device is None:
            device = "cuda" if torch.cuda.is_available() else "cpu"
        self.device = torch.device(device)
        self.model.to(device)

    def _mean_pooling(self, model_output, attention_mask):
        token_embeddings = model_output[0] #First element of model_output contains all token embeddings
        input_mask_expanded = attention_mask.unsqueeze(-1).expand(token_embeddings.size()).float()
        return torch.sum(token_embeddings * input_mask_expanded, 1) / torch.clamp(input_mask_expanded.sum(1), min=1e-9)

    @torch.no_grad()
    def encode(self, sentences, batch_size=8):
        all_embeddings = []
        iterator = range(0, len(sentences), batch_size)
        for batch_idx in iterator:
            batch = sentences[batch_idx:batch_idx + batch_size]

            # encoded_input = self.tokenizer.batch_encode_plus(batch, padding="longest", 
            #                                truncation=True, return_tensors="pt").to(self.device)
            encoded_input = self.tokenizer.batch_encode_plus(batch, padding="max_length", max_length=512,
                                           truncation=True, return_tensors="pt").to(self.device)
            model_output = self.model(**encoded_input)
            sentence_embeddings = self._mean_pooling(model_output, encoded_input["attention_mask"]).to('cpu')

            all_embeddings.extend(sentence_embeddings)

        # return torch.stack(all_embeddings).numpy()
        return torch.stack(all_embeddings)



model = SentenceBertJapanese("sonoisa/sentence-bert-base-ja-mean-tokens-v2")

def cos_sim(v1, v2):
    return np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2))

def calculate_euclidean_distance(vec1, vec2):
    """
    2つのベクトル間のユークリッド距離を計算する。
    :return: ベクトル間のユークリッド距離
    """
    distance = np.linalg.norm(vec1 - vec2)
    # ユークリッド距離が0の場合、類似度は最大とする
    if distance == 0:
        return np.array([1.0], dtype=np.float32)
    else:
        # 類似度は距離の逆数とする（距離が小さいほど類似度が高い）
        similarity = 1 / (1 + distance)
        return np.array([similarity], dtype=np.float32)

def calc(input_content, contents):

    # 入力文章のベクトル化
    input_vec = model.encode([input_content])

    content_vecs = model.encode(contents)
    content_vec_average = sum([content_vec.numpy() for content_vec in content_vecs]) / len(content_vecs)

    # # 類似度計算
    similarities = [cos_sim(input_vec, vec) for vec in content_vecs]

    return sum(similarities)/len(similarities)

def read_data_from_excel(file_path, cells):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    data = []
    for cell_name in cells:
        cell = sheet[cell_name]
        value = get_merged_cell_value(sheet, cell)
        if value is not None:
            # テキストデータをリストに追加
            data.append(value)
    return data

def get_merged_cell_value(sheet, cell):
    for merged_cell in sheet.merged_cells.ranges:
        if cell.coordinate in merged_cell:
            # 結合されたセルの左上のセルの値を返す
            return sheet[merged_cell.min_row][merged_cell.min_col - 1].value
    # 結合されていないセルの場合、そのセルの値を返す
    return cell.value


def main():
    # Excelファイルのパス
    path_contents = './close/data/論文生成結果標準偏差.xlsx'

    # 読み込むセル（2つのグループ）
    cells_to_read_c = ['C17', 'C40', 'C63', 'C86', 'C109']
    cells_to_read_o = ['O17', 'O40', 'O63', 'O86', 'O109']

    # ワークブックを開く
    workbook = openpyxl.load_workbook(path_contents)

    # ワークブック内の全てのシートに対して処理を行う
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # C列のデータを読み込んで計算
        contents_c = read_data_from_sheet(sheet, cells_to_read_c)
        sim_c = calculate_sim(contents_c)

        if len(sim_c) >= 2:
            stdev_c = statistics.stdev(sim_c)
            print(stdev_c)
            sheet['C12'] = stdev_c  # C列の結果をC12に書き込む

        # O列のデータを読み込んで計算
        contents_o = read_data_from_sheet(sheet, cells_to_read_o)
        sim_o = calculate_sim(contents_o)

        if len(sim_o) >= 2:
            stdev_o = statistics.stdev(sim_o)
            print(stdev_o)
            sheet['O12'] = stdev_o  # O列の結果をO12に書き込む

    # 変更を保存
    workbook.save(path_contents)

def calculate_sim(contents):
    sim = []
    for i in range(len(contents)):
        new_list = [x for j, x in enumerate(contents) if j != i]
        sim.append(calc(contents[i], new_list)[0])
    return sim

def read_data_from_sheet(sheet, cells):
    data = []
    for cell_name in cells:
        cell = sheet[cell_name]
        value = get_merged_cell_value(sheet, cell)
        if value is not None:
            data.append(value)
    return data

# get_merged_cell_value 関数は以前と同じ




if __name__ == "__main__":
    main()